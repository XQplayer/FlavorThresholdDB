from pathlib import Path

from openpyxl import load_workbook


PATH = Path(__file__).resolve().parents[2] / "data" / "work" / "批量CAS_阈值_风味描述查询.xlsx"
ws = load_workbook(PATH, data_only=True)["Summary"]
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
idx = {header: i + 1 for i, header in enumerate(headers)}

fields = ["CAS", "英文名", "本地阈值库介质", "本地阈值条数", "FEMA风味描述", "匹配备注"]
print("\t".join(fields))
for row in range(2, ws.max_row + 1):
    values = []
    for field in fields:
        value = str(ws.cell(row, idx[field]).value or "").replace("\n", "; ")
        values.append(value[:220])
    print("\t".join(values))
