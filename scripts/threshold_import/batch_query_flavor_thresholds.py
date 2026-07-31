from __future__ import annotations

import html
import json
import re
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CAS_LIST = [
    "105-37-3",
    "556-67-2",
    "105-54-4",
    "29949-66-4",
    "78-83-1",
    "123-92-2",
    "95-47-6",
    "541-02-6",
    "123-51-3",
    "123-66-0",
]

PROJECT_ROOT = Path(__file__).resolve().parents[2]
AROMA_DB = PROJECT_ROOT / "frontend" / "public" / "aroma_data_merged.json"
LOCAL_LIBRARY = Path(r"F:\1GSAU2022.9-2025.6\韩兴权硕士毕业材料\0试验数据\1试验相关数据\方法\解谱方法\解谱库2024.4.8更新-韩兴权.xlsx")
FEMA_CACHE = PROJECT_ROOT / "fema_flavor_cache.json"
OUTPUT = PROJECT_ROOT / "data" / "work" / "批量CAS_阈值_风味描述查询.xlsx"

BASE_URL = "https://www.femaflavor.org"


def norm(value) -> str:
    return "" if value is None else str(value).strip()


def compact_list(values: list[str]) -> str:
    seen = []
    for value in values:
        value = norm(value)
        if value and value not in seen:
            seen.append(value)
    return "\n".join(seen)


def load_aroma_db() -> dict[str, list[dict]]:
    data = json.loads(AROMA_DB.read_text(encoding="utf-8"))
    out: dict[str, list[dict]] = {}
    for item in data:
        cas = norm(item.get("cas"))
        if cas:
            out.setdefault(cas, []).append(item)
    return out


def load_local_library() -> dict[str, list[dict]]:
    if not LOCAL_LIBRARY.exists():
        return {}
    wb = load_workbook(LOCAL_LIBRARY, data_only=True, read_only=True)
    records: dict[str, list[dict]] = {}
    for sheet_name in ["黄酒新增", "韩兴权库", "Sheet1"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        idx = {norm(header): i + 1 for i, header in enumerate(headers) if header is not None}
        if "CAS#" not in idx:
            continue
        for row in range(2, ws.max_row + 1):
            cas = norm(ws.cell(row, idx["CAS#"]).value)
            if not cas:
                continue
            rec = {
                "sheet": sheet_name,
                "CNC": ws.cell(row, idx["CNC"]).value if "CNC" in idx else None,
                "CNE": ws.cell(row, idx["CNE"]).value if "CNE" in idx else None,
                "Polar": ws.cell(row, idx["Polar"]).value if "Polar" in idx else None,
                "CC": ws.cell(row, idx["CC"]).value if "CC" in idx else None,
                "空气阈值": ws.cell(row, idx["空气阈值"]).value if "空气阈值" in idx else None,
                "水中阈值": ws.cell(row, idx.get("水中阈值 ", idx.get("水中阈值 \n", 0))).value if ("水中阈值 " in idx or "水中阈值 \n" in idx) else None,
                "其它阈值": ws.cell(row, idx["其它阈值"]).value if "其它阈值" in idx else None,
                "Flavor profile": ws.cell(row, idx["Flavor profile"]).value if "Flavor profile" in idx else None,
                "特征描述": ws.cell(row, idx["特征描述"]).value if "特征描述" in idx else None,
            }
            records.setdefault(cas, []).append(rec)
    return records


def load_fema_cache() -> dict:
    if not FEMA_CACHE.exists():
        return {}
    try:
        return json.loads(FEMA_CACHE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_fema_cache(cache: dict) -> None:
    FEMA_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_text(url: str) -> str:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=25) as resp:
        return resp.read().decode("utf-8", errors="replace")


def strip_tags(value: str) -> str:
    value = re.sub(r"<br\s*/?>", "\n", value, flags=re.I)
    value = re.sub(r"</p\s*>", "\n", value, flags=re.I)
    value = re.sub(r"<[^>]+>", " ", value)
    value = html.unescape(value)
    value = re.sub(r"[ \t\r\f\v]+", " ", value)
    value = re.sub(r"\n\s+", "\n", value)
    return value.strip()


def first_match(pattern: str, text: str, flags: int = re.S | re.I) -> str:
    match = re.search(pattern, text, flags)
    return match.group(1).strip() if match else ""


def query_fema(query: str, cache: dict) -> dict:
    key = query.lower()
    if key in cache:
        return cache[key]

    search_url = f"{BASE_URL}/flavor-library/search?fulltext={quote(query)}"
    try:
        search_html = fetch_text(search_url)
        href = first_match(r'<div class="views-row ingredient-item">\s*<a href="([^"]+)"', search_html)
        if not href:
            result = {"found": False, "query": query, "error": "No FEMA result"}
        else:
            detail_url = href if href.startswith("http") else BASE_URL + href
            detail_html = fetch_text(detail_url)
            name_match = re.search(r"<h1[^>]*>\s*([^<]*?)<strong>\s*<span>(.*?)</span>", detail_html, re.S | re.I)
            fema_number = strip_tags(name_match.group(1)) if name_match else ""
            name = strip_tags(name_match.group(2)) if name_match else ""
            cas = strip_tags(first_match(r'field--name-field-cas[\s\S]*?<div class="field__item">([\s\S]*?)</div>', detail_html))
            flavor_profile = strip_tags(first_match(r'<div class="field[^"]*field--name-field-flavor-profile[^"]*"[^>]*>([\s\S]*?)</div>', detail_html))
            jecfa = strip_tags(first_match(r'field--name-field-jecfa-number[\s\S]*?<div class="field__item">([\s\S]*?)</div>', detail_html))
            result = {
                "found": bool(name or flavor_profile),
                "query": query,
                "name": name,
                "cas": cas,
                "fema_number": fema_number,
                "jecfa_number": jecfa,
                "flavor_profile": flavor_profile,
                "url": detail_url,
                "source": "FEMA Flavor Library",
            }
    except Exception as exc:
        result = {"found": False, "query": query, "error": str(exc)}

    cache[key] = result
    save_fema_cache(cache)
    return result


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, ws.max_column + 1):
        width = len(str(ws.cell(1, col).value or ""))
        for row in range(1, min(ws.max_row, 80) + 1):
            width = max(width, min(80, len(str(ws.cell(row, col).value or ""))))
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 45)


def main() -> None:
    aroma = load_aroma_db()
    local = load_local_library()
    fema_cache = load_fema_cache()

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary_headers = [
        "CAS",
        "英文名",
        "中文名",
        "本地阈值库介质",
        "本地阈值条数",
        "解谱库类别",
        "解谱库Polar",
        "解谱库空气阈值",
        "解谱库水中阈值",
        "解谱库其它阈值",
        "FEMA名称",
        "FEMA编号",
        "FEMA风味描述",
        "本地/解谱库风味描述",
        "FEMA链接",
        "匹配备注",
    ]
    summary.append(summary_headers)

    detail = wb.create_sheet("Threshold_Details")
    detail_headers = ["CAS", "英文名", "中文名", "介质", "单位", "阈值记录"]
    detail.append(detail_headers)

    library_sheet = wb.create_sheet("Local_Library")
    library_headers = ["CAS", "来源Sheet", "英文名", "中文名", "Polar", "类别", "空气阈值", "水中阈值", "其它阈值", "Flavor profile", "特征描述"]
    library_sheet.append(library_headers)

    fema_sheet = wb.create_sheet("FEMA")
    fema_headers = ["CAS", "Found", "FEMA名称", "FEMA编号", "JECFA编号", "FEMA风味描述", "URL", "Error"]
    fema_sheet.append(fema_headers)

    for cas in CAS_LIST:
        aroma_records = aroma.get(cas, [])
        local_records = local.get(cas, [])
        fema = query_fema(cas, fema_cache)

        english = compact_list(
            [r.get("english_name", "") for r in aroma_records]
            + [norm(r.get("CNE")) for r in local_records]
            + [norm(fema.get("name"))]
        )
        chinese = compact_list(
            [r.get("chinese_name", "") for r in aroma_records]
            + [norm(r.get("CNC")) for r in local_records]
        )
        media = compact_list([r.get("medium", "") for r in aroma_records])
        threshold_count = sum(len(r.get("threshold_data") or []) for r in aroma_records)
        local_desc = compact_list(
            [desc for r in aroma_records for desc in (r.get("flavor_desc_cn") or [])]
            + [desc for r in aroma_records for desc in (r.get("flavor_desc") or [])]
            + [norm(r.get("Flavor profile")) for r in local_records]
            + [norm(r.get("特征描述")) for r in local_records]
        )
        notes = []
        if not aroma_records:
            notes.append("本地阈值库未匹配")
        if not local_records:
            notes.append("解谱库未匹配")
        if not fema.get("found"):
            notes.append("FEMA未匹配")

        summary.append(
            [
                cas,
                english,
                chinese,
                media,
                threshold_count,
                compact_list([norm(r.get("CC")) for r in local_records]),
                compact_list([norm(r.get("Polar")) for r in local_records]),
                compact_list([norm(r.get("空气阈值")) for r in local_records]),
                compact_list([norm(r.get("水中阈值")) for r in local_records]),
                compact_list([norm(r.get("其它阈值")) for r in local_records]),
                fema.get("name", ""),
                fema.get("fema_number", ""),
                fema.get("flavor_profile", ""),
                local_desc,
                fema.get("url", ""),
                "；".join(notes),
            ]
        )

        for record in aroma_records:
            unit = "mg/m³" if record.get("medium") == "空气" else "mg/kg"
            detail.append(
                [
                    cas,
                    record.get("english_name", ""),
                    record.get("chinese_name", ""),
                    record.get("medium", ""),
                    unit,
                    "\n".join(record.get("threshold_data") or []),
                ]
            )

        for record in local_records:
            library_sheet.append(
                [
                    cas,
                    record.get("sheet", ""),
                    record.get("CNE", ""),
                    record.get("CNC", ""),
                    record.get("Polar", ""),
                    record.get("CC", ""),
                    record.get("空气阈值", ""),
                    record.get("水中阈值", ""),
                    record.get("其它阈值", ""),
                    record.get("Flavor profile", ""),
                    record.get("特征描述", ""),
                ]
            )

        fema_sheet.append(
            [
                cas,
                fema.get("found", False),
                fema.get("name", ""),
                fema.get("fema_number", ""),
                fema.get("jecfa_number", ""),
                fema.get("flavor_profile", ""),
                fema.get("url", ""),
                fema.get("error", ""),
            ]
        )

    for ws in wb.worksheets:
        style_sheet(ws)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"output={OUTPUT}")
    print(f"cas_count={len(CAS_LIST)}")


if __name__ == "__main__":
    main()
